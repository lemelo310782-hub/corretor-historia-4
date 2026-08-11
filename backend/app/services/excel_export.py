"""
Gera o relatório da turma em Excel:

    Aluno | Critério 1 | Critério 2 | ... | Nota final | Status
"""
import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import EXPORTS_DIR

COR_CABECALHO = "0F2038"  # azul escuro acadêmico, sem "#" (openpyxl espera hex puro)


def _slug(texto: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "_", sem_acento).strip("_").lower()[:60] or "atividade"


def gerar_excel_turma(atividade) -> Path:
    """
    `atividade` precisa ter `.correcoes` (cada uma com `.criterios` e
    `.aluno`) já carregáveis. Retorna o caminho do .xlsx gerado em
    EXPORTS_DIR.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório da turma"

    # União dos nomes de critério, na ordem em que aparecem pela primeira vez
    nomes_criterios: list[str] = []
    for correcao in atividade.correcoes:
        for c in correcao.criterios:
            if c.nome_criterio not in nomes_criterios:
                nomes_criterios.append(c.nome_criterio)

    cabecalho = ["Aluno"] + nomes_criterios + ["Nota final", "Status"]
    ws.append(cabecalho)

    fundo_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    for celula in ws[1]:
        celula.fill = fundo_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for correcao in atividade.correcoes:
        nome = correcao.aluno.nome if correcao.aluno else (correcao.nome_detectado or f"Ficha #{correcao.id}")
        pontuacoes_por_criterio = {c.nome_criterio: c.pontuacao_obtida for c in correcao.criterios}

        status_valor = correcao.status.value if hasattr(correcao.status, "value") else correcao.status
        linha = (
            [nome]
            + [pontuacoes_por_criterio.get(nome_c, "") for nome_c in nomes_criterios]
            + [correcao.nota_final if correcao.nota_final is not None else "", status_valor]
        )
        ws.append(linha)

    # Auto-ajuste simples de largura de coluna
    for indice_coluna, _ in enumerate(cabecalho, start=1):
        letra = get_column_letter(indice_coluna)
        maior_valor = max(
            (len(str(celula.value)) for celula in ws[letra] if celula.value is not None),
            default=10,
        )
        ws.column_dimensions[letra].width = min(maior_valor + 3, 45)

    ws.freeze_panes = "A2"

    caminho = EXPORTS_DIR / f"relatorio_turma_atividade_{atividade.id}_{_slug(atividade.titulo)}.xlsx"
    wb.save(caminho)
    return caminho
